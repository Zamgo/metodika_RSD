"""
Convert RACI Excel matrix to individual Obsidian MD files.

Reads: (šablona) Příloha_1_Matice_informačního_managementu_(RACI).xlsx
Writes: ../07_RACI_cinnosti/*.md  +  ../Cinnosti.base
"""

import openpyxl
import os
import re
import shutil
import unicodedata

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
VAULT_ROOT = os.path.dirname(SCRIPT_DIR)
OUTPUT_DIR = os.path.join(VAULT_ROOT, "07_RACI_cinnosti")
EXCEL_NAME = "(šablona) Příloha_1_Matice_informačního_managementu_(RACI).xlsx"

ROLE_COLS = {
    4: "poverujici strana",      # col E
    5: "vedouci poverena strana", # col F
    6: "poverena strana",         # col G
    8: "spravce stavby",          # col I
    9: "bim koordinator",         # col J
}

RACI_KEYS = {
    4: "raci_poverejici",
    5: "raci_vedouci_poverena",
    6: "raci_poverena",
    8: "raci_spravce_stavby",
    9: "raci_bim_koordinator",
}

ISO_FAZE_MAP = {
    "1": ("5.1", ["priprava"]),
    "2": ("5.2", ["priprava"]),
    "3": ("5.3", ["priprava"]),
    "4": ("5.4", ["priprava"]),
    "5": ("5.5", ["realizace"]),
    "6": ("5.6", ["realizace"]),
    "7": ("5.7", ["realizace"]),
    "8": ("5.8", ["provoz_a_udrzba"]),
}


def slugify(text: str) -> str:
    """Remove diacritics and produce a filesystem-safe slug."""
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_text = nfkd.encode("ascii", "ignore").decode("ascii")
    ascii_text = re.sub(r"[^\w\s-]", "", ascii_text).strip()
    return re.sub(r"[\s]+", " ", ascii_text)[:80]


def derive_roles(raci_values: dict) -> list:
    """Derive role list from non-empty RACI values containing R or A."""
    roles = []
    for col_idx, role_name in ROLE_COLS.items():
        val = raci_values.get(col_idx, "")
        if val and ("R" in val or "A" in val):
            roles.append(role_name)
    return roles


def derive_zdroj_typ(zdroj: str) -> str:
    if not zdroj or zdroj.lower() == "n/a":
        return "interni_metodika"
    zdroj_lower = zdroj.lower()
    if "19650" in zdroj_lower or "iso" in zdroj_lower:
        return "iso_19650"
    if "fidic" in zdroj_lower:
        return "fidic"
    if "eir" in zdroj_lower:
        return "eir"
    return "interni_metodika"


def get_parent_id(id_str: str) -> str:
    parts = id_str.split(".")
    if len(parts) <= 1:
        return ""
    return ".".join(parts[:-1])


def get_section_id(id_str: str) -> str:
    return id_str.split(".")[0]


def yaml_list(items: list) -> str:
    if not items:
        return "[]"
    return "[" + ", ".join(items) + "]"


def yaml_str(val: str) -> str:
    if not val:
        return '""'
    if any(c in val for c in ':#{}[]|>&*!?,'):
        return '"' + val.replace('"', '\\"') + '"'
    return '"' + val + '"'


def generate_md(row_data: dict, subsections: dict) -> str:
    """Generate full MD file content for a single RACI activity."""
    fm_lines = ["---"]
    fm_lines.append(f"title: {yaml_str(row_data['title'])}")
    fm_lines.append("typ: raci_cinnost")
    fm_lines.append(f"oznaceni: {yaml_str(row_data['id'])}")
    fm_lines.append(f"sekce: {yaml_str(row_data['parent_id'])}")
    fm_lines.append(f"iso_faze: {yaml_str(row_data['iso_faze'])}")
    fm_lines.append(f"popis: {yaml_str(row_data['desc'])}")
    fm_lines.append(f"zdroj: {yaml_str(row_data['zdroj'])}")
    fm_lines.append(f"zdroj_typ: {row_data['zdroj_typ']}")
    fm_lines.append(f"faze: {yaml_list(row_data['faze'])}")
    fm_lines.append(f"role: {yaml_list(row_data['roles'])}")

    for col_idx, key in RACI_KEYS.items():
        val = row_data["raci"].get(col_idx, "")
        fm_lines.append(f"{key}: {yaml_str(val)}")

    fm_lines.append("workflow: []")
    fm_lines.append("stav: draft")
    fm_lines.append(f"tags: {yaml_list(row_data['tags'])}")
    fm_lines.append("---")
    fm_lines.append("")

    body_lines = []
    if row_data["desc"]:
        body_lines.append("## Popis")
        body_lines.append("")
        body_lines.append(row_data["desc"])
        body_lines.append("")

    body_lines.append("## Zdroj")
    body_lines.append("")
    body_lines.append(row_data["zdroj"] if row_data["zdroj"] else "—")
    body_lines.append("")

    parent_title = subsections.get(row_data["parent_id"], "")
    if parent_title:
        body_lines.append("## Návaznosti")
        body_lines.append("")
        body_lines.append(
            f"- Nadřazená sekce: {row_data['parent_id']} – {parent_title}"
        )
        body_lines.append("")

    return "\n".join(fm_lines + body_lines)


def generate_base() -> str:
    """Generate the Cinnosti.base YAML content."""
    return """filters:
  or:
    - file.inFolder("02_Oblasti správy informací")
    - file.inFolder("07_RACI_cinnosti")
properties:
  oznaceni:
    displayName: "ID"
  zdroj:
    displayName: "Zdroj"
  zdroj_typ:
    displayName: "Typ zdroje"
  raci_poverejici:
    displayName: "Objednatel"
  raci_vedouci_poverena:
    displayName: "Zhotovitel"
  raci_poverena:
    displayName: "Podzhotovitel"
  raci_spravce_stavby:
    displayName: "Správce stavby"
  raci_bim_koordinator:
    displayName: "BIM koord."
views:
  - type: table
    name: "Všechny činnosti"
    order:
      - file.name
      - typ
      - faze
      - role
      - zdroj_typ
      - zdroj
      - raci_poverejici
      - raci_vedouci_poverena
      - raci_spravce_stavby
      - raci_bim_koordinator
      - stav
"""


def main():
    excel_path = os.path.join(SCRIPT_DIR, EXCEL_NAME)
    temp_path = os.path.join(os.environ.get("TEMP", "/tmp"), "raci_convert.xlsx")

    try:
        shutil.copy2(excel_path, temp_path)
    except PermissionError:
        if os.path.exists(temp_path):
            print("Source locked by OneDrive, using existing temp copy.")
        else:
            raise

    wb = openpyxl.load_workbook(temp_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    all_rows = []
    subsections = {}  # id -> title for depth-2 rows

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        id_val = row[0]
        if id_val is None or str(id_val).strip() == "":
            continue

        id_str = str(id_val).strip()
        # Handle numeric IDs that might come as floats (e.g. 1.0 -> "1")
        if id_str.endswith(".0") and "." not in id_str[:-2]:
            id_str = id_str[:-2]

        depth = len(id_str.split("."))
        title = str(row[1]).strip() if row[1] else ""
        desc = str(row[2]).strip() if row[2] else ""
        zdroj = str(row[3]).strip() if row[3] else ""

        raci = {}
        for col_idx in RACI_KEYS:
            val = row[col_idx]
            raci[col_idx] = str(val).strip() if val else ""

        if depth == 2:
            subsections[id_str] = title

        section = get_section_id(id_str)
        iso_faze, faze = ISO_FAZE_MAP.get(section, ("", []))

        all_rows.append({
            "id": id_str,
            "depth": depth,
            "title": title,
            "desc": desc,
            "zdroj": zdroj,
            "zdroj_typ": derive_zdroj_typ(zdroj),
            "raci": raci,
            "roles": derive_roles(raci),
            "iso_faze": iso_faze,
            "faze": faze,
            "parent_id": get_parent_id(id_str),
            "tags": ["raci", "iso_19650"],
        })

    # Generate MD files only for leaf tasks (depth >= 3)
    # and subsections (depth == 2) that have their own RACI values
    count = 0
    for row_data in all_rows:
        if row_data["depth"] < 2:
            continue
        # Skip depth-2 rows that have no RACI values at all
        if row_data["depth"] == 2:
            has_raci = any(v for v in row_data["raci"].values())
            if not has_raci:
                continue

        slug = slugify(row_data["title"])
        if not slug:
            slug = row_data["id"].replace(".", "_")
        filename = f"{row_data['id']} - {slug}.md"
        filepath = os.path.join(OUTPUT_DIR, filename)

        content = generate_md(row_data, subsections)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)
        count += 1

    # Generate Cinnosti.base in vault root
    base_path = os.path.join(VAULT_ROOT, "Cinnosti.base")
    with open(base_path, "w", encoding="utf-8") as f:
        f.write(generate_base())

    print(f"Generated {count} MD files in {OUTPUT_DIR}")
    print(f"Generated Cinnosti.base in {VAULT_ROOT}")


if __name__ == "__main__":
    main()
